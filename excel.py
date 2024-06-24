from openpyxl import Workbook
from openpyxl import load_workbook
from openpyxl import load_workbook
from fpdf import FPDF

def postExcel(listaUsuarios):
    # Criando um novo arquivo Excel
    newExcel = Workbook()

    # Selecionando a planilha ativa (por padrão, há uma planilha chamada 'Sheet')
    sheet = newExcel.active
    
    # Adicionando dados à planilha
    sheet['A1'] = 'Nome'
    sheet['B1'] = 'Idade'
    sheet['C1'] = 'Peso'
    sheet['D1'] = 'Têm apelido?'
    sheet['E1'] = 'Apelido'
    sheet['F1'] = 'Data de Nascimento'
    
    # Adicionando dados dos usuários à planilha
    for idx, usuario in enumerate(listaUsuarios, start=2):  # começa a partir da linha 2
        sheet[f'A{idx}'] = usuario['nome']
        sheet[f'B{idx}'] = usuario['idade']
        sheet[f'C{idx}'] = usuario['peso']
        sheet[f'D{idx}'] = 'Sim' if usuario['apelidoBool'] else 'Não'
        sheet[f'E{idx}'] = usuario['apelido'] if usuario['apelido'] else ''
        sheet[f'F{idx}'] = usuario['dataNascimento']
        
        
    # Salvando o arquivo Excel
    usuarios = 'tabelaUsuarios.xlsx'
    newExcel.save(usuarios)
    print(f'Arquivo Excel "{usuarios}" criado com sucesso.')

    # Fechando o arquivo Excel criado
    newExcel.close()

    # Abrindo o arquivo Excel recém-criado para leitura
    newExcel_aberto = load_workbook(usuarios)

    # Acessando a planilha ativa
    sheet_aberta = newExcel_aberto.active

    # Exibindo os dados da planilha aberta
    print('\nDados da planilha aberta:')
    for row in sheet_aberta.iter_rows(values_only=True):
        print(row)

    # Fechando o arquivo Excel aberto
    newExcel_aberto.close()
    # Exemplo de uso
    excel_para_pdf('tabelaUsuarios.xlsx', 'tabelaUsuarios.pdf')
    
    
def excel_para_pdf(excel, novoPdf):
    # Carregar o arquivo Excel
    arquvi_excel = load_workbook(excel)
    planilha_excel = arquvi_excel.active

    # Configurar o PDF
    pdf = FPDF()
    pdf.add_page()
    pdf.set_font('Arial', size=12)

    # Iterar pelas linhas do Excel e adicionar ao PDF
    for row in planilha_excel.iter_rows():
        for cell in row:
            pdf.cell(200, 10, str(cell.value), ln=True)

    # Salvar o PDF
    pdf.output(novoPdf)
